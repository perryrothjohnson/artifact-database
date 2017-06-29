#!/usr/bin/env python
import os
import glob
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copy

# define 4 types of spreadsheets
# list_of_sheet_types = ['accounting']
list_of_sheet_types = ['accounting', 'dennis', 'exdev', 'restoration_conservation']

for sheet_type in list_of_sheet_types:
    # go to directory with CSV file of exported artifacts
    os.chdir('/var/www/html/support/export_spreadsheets/data/' + sheet_type)

    # grab the most recent csv file
    list_of_csv_files = glob.glob(os.path.join('.', '*.csv'))
    list_of_csv_files.sort()
    csvfile = list_of_csv_files[-1]  # last file in sorted list should be most recent

    wb = Workbook()
    ws = wb.active

    # read in csvfile with pandas, then sort by first four columns
    raw_data = pd.read_csv(csvfile)
    if sheet_type == 'accounting':
        sorted_data = raw_data.sort_values(by=['Supergroup', 'Object name'])
    if sheet_type == 'dennis':
        sorted_data = raw_data.sort_values(by=['Supergroup','Group','Object name','Object ID number'])
    if sheet_type == 'exdev':
        sorted_data = raw_data.sort_values(by=['Phase III exhibit code', 'Group', 'Object name', 'Object ID number'])
    if sheet_type == 'restoration_conservation':
        sorted_data = raw_data.sort_values(by=['Group', 'Object name', 'Object ID number'])

    # export data from pandas to openpyxl
    for r in dataframe_to_rows(sorted_data, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # freeze panes: header row
    # ref: https://stackoverflow.com/questions/25588918/how-to-freeze-entire-header-row-in-openpyxl
    if sheet_type == 'accounting':
        pivot = ws['B2']
    if sheet_type == 'dennis':
        pivot = ws['D2']
    if sheet_type == 'exdev':
        pivot = ws['D2']
    if sheet_type == 'restoration_conservation':
        pivot = ws['C2']
    ws.freeze_panes = pivot

    # set the column dimensions
    if sheet_type == 'accounting':
        ws.column_dimensions["A"].width = 25.0			# Supergroup
	ws.column_dimensions["B"].width = 50.0                  # Object name
        ws.column_dimensions["C"].width =  8.2                  # Quantity
        ws.column_dimensions["D"].width = 34.0                  # Donor...party
        ws.column_dimensions["E"].width = 20.0                  # Transfer date
        ws.column_dimensions["F"].width = 20.0                  # Delivery date
        ws.column_dimensions["G"].width =  9.25                 # On display?
        ws.column_dimensions["H"].width = 70.0                  # Current location
        ws.column_dimensions["I"].width = 15.75                 # Item value
        ws.column_dimensions["J"].width = 14.29                 # Insurance value
        ws.column_dimensions["K"].width = 20.0                  # Insurance method
    if sheet_type == 'dennis':
        ws.column_dimensions["A"].width = 25.0                  # Supergroup
        ws.column_dimensions["B"].width = 32.0                  # Group
        ws.column_dimensions["C"].width = 50.0                  # Object name
        ws.column_dimensions["D"].width = 16.0                  # Object ID number
        ws.column_dimensions["E"].width =  8.2                  # Quantity
        ws.column_dimensions["F"].width = 21.14                 # Part number
        ws.column_dimensions["G"].width = 23.57                 # Serial number
        ws.column_dimensions["H"].width = 34.0                  # Donor...party
        ws.column_dimensions["I"].width = 15.63                 # Acquisition method
        ws.column_dimensions["J"].width = 27.0                  # Object lot ID
        ws.column_dimensions["K"].width =  9.86                 # DGS paid
        ws.column_dimensions["L"].width = 20.0                  # Transfer date
        ws.column_dimensions["M"].width = 20.0                  # Delivery date
        ws.column_dimensions["N"].width =  9.25                 # On display?
        ws.column_dimensions["O"].width = 70.0                  # Current location
        ws.column_dimensions["P"].width = 12.43                 # Accessioned?
        ws.column_dimensions["Q"].width = 10.0                  # ITAR status
        ws.column_dimensions["R"].width = 14.86                 # ITAR code
        ws.column_dimensions["S"].width = 19.14                 # Owned by/loaned to
        ws.column_dimensions["T"].width = 15.75                 # Item value
        ws.column_dimensions["U"].width = 14.29                 # Insurance value
        ws.column_dimensions["V"].width = 20.0                  # Insurance method
        ws.column_dimensions["W"].width = 99.0                  # Notes
    if sheet_type == 'exdev':
        ws.column_dimensions["A"].width = 17.0                  # Phase III exhibit code
        ws.column_dimensions["B"].width = 32.0                  # Group
        ws.column_dimensions["C"].width = 50.0                  # Object name
        ws.column_dimensions["D"].width = 50.0                  # Alt. object name
        ws.column_dimensions["E"].width = 16.0                  # Object ID number
        ws.column_dimensions["F"].width =  8.2                  # Quantity
        ws.column_dimensions["G"].width = 34.0                  # Donor...party
        ws.column_dimensions["H"].width = 34.0                  # Donor attribution
        ws.column_dimensions["I"].width = 15.63                 # Acquisition method
        ws.column_dimensions["J"].width =  7.75                 # Loan type
        ws.column_dimensions["K"].width = 15.0                  # Loan return date
        ws.column_dimensions["L"].width =  9.25                 # On display?
        ws.column_dimensions["M"].width = 70.0                  # Current location
        ws.column_dimensions["N"].width = 25.0                  # Intended display method
        ws.column_dimensions["O"].width =  8.5                  # Length
        ws.column_dimensions["P"].width =  8.5                  # Width
        ws.column_dimensions["Q"].width =  8.5                  # Height
        ws.column_dimensions["R"].width =  8.5                  # Weight
        ws.column_dimensions["S"].width = 19.14                 # Owned by/loaned to
        ws.column_dimensions["T"].width = 16.25                 # Date of manufacture
    if sheet_type == 'restoration_conservation':
        ws.column_dimensions["A"].width = 32.0                  # Group
        ws.column_dimensions["B"].width = 50.0                  # Object name
        ws.column_dimensions["C"].width = 50.0                  # Alt. object name
        ws.column_dimensions["D"].width = 16.0                  # Object ID number
        ws.column_dimensions["E"].width =  8.2                  # Quantity
        ws.column_dimensions["F"].width = 21.14                 # Part number
        ws.column_dimensions["G"].width = 23.57                 # Serial number
        ws.column_dimensions["H"].width = 50.0                  # Other ID numbers
        ws.column_dimensions["I"].width = 34.0                  # Donor...party
        ws.column_dimensions["J"].width = 17.0                  # Contact
        ws.column_dimensions["K"].width = 15.63                 # Acquisition method
        ws.column_dimensions["L"].width =  7.75                 # Loan type
        ws.column_dimensions["M"].width = 17.0                  # Loan ID
        ws.column_dimensions["N"].width = 15.0                  # Loan in date
        ws.column_dimensions["O"].width = 15.0                  # Loan out date
        ws.column_dimensions["P"].width = 15.0                  # Loan return date
        ws.column_dimensions["Q"].width = 34.0                  # Lender/Loaned to
        ws.column_dimensions["R"].width = 27.0                  # Object lot ID
        ws.column_dimensions["S"].width = 20.0                  # Transfer date
        ws.column_dimensions["T"].width = 20.0                  # Delivery date
        ws.column_dimensions["U"].width =  9.25                 # On display?
        ws.column_dimensions["V"].width = 70.0                  # Current location
        ws.column_dimensions["W"].width = 25.0                  # Intended display method
        ws.column_dimensions["X"].width = 12.43                 # Accessioned?
        ws.column_dimensions["Y"].width = 12.43                 # Deaccessioned?
        ws.column_dimensions["Z"].width = 10.0                  # ITAR status
        ws.column_dimensions["AA"].width = 14.86                # ITAR code
        ws.column_dimensions["AB"].width =  8.5                 # Length
        ws.column_dimensions["AC"].width =  8.5                 # Width
        ws.column_dimensions["AD"].width =  8.5                 # Height
        ws.column_dimensions["AE"].width =  8.5                 # Weight
        ws.column_dimensions["AF"].width = 19.14                # Owned by/loaned to
        ws.column_dimensions["AG"].width = 15.75                # Item value
        ws.column_dimensions["AH"].width = 14.29                # Insurance value
        ws.column_dimensions["AI"].width = 20.0                 # Insurance method
        ws.column_dimensions["AJ"].width =  8.0                 # Condition
        ws.column_dimensions["AK"].width = 70.0                 # Object materials
        ws.column_dimensions["AL"].width = 50.0                 # Special handling/maintenance
        ws.column_dimensions["AM"].width = 50.0                 # Cons. strategy - Cleaning
        ws.column_dimensions["AN"].width = 50.0                 # Cons. strategy - Storage
        ws.column_dimensions["AO"].width = 50.0                 # Cons. strategy - Display
        ws.column_dimensions["AP"].width = 50.0                 # Environmental requirements
        ws.column_dimensions["AQ"].width = 15.63                # Est. restoration cost
        ws.column_dimensions["AR"].width = 20.25                # Source of restoration cost
        ws.column_dimensions["AS"].width = 16.25                # Date of manufacture
        ws.column_dimensions["AT"].width = 50.0                 # Object function
        ws.column_dimensions["AU"].width = 50.0                 # Object description
        ws.column_dimensions["AV"].width = 50.0                 # Historical significance
        ws.column_dimensions["AW"].width = 50.0                 # Pedigree
        ws.column_dimensions["AX"].width = 99.0                 # Notes

    # trim off the .csv extension and save as .xlsx file
    xlsxfile = csvfile[:-4] + '.xlsx'
    wb.save(xlsxfile)

    # copy this latest xlsx file to Dropbox
    copy(xlsxfile, '/home/perry/Dropbox/CSC In-house/Artifact Documents/!Donations and Acquisitions/' + sheet_type)

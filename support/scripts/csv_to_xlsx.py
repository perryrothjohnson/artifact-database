#!/usr/bin/env python

import os
import glob
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copy

# go to directory with CSV file of exported artifacts
os.chdir('/var/www/html/support/dennis_export/data')

# grab the most recent csv file
list_of_csv_files = glob.glob(os.path.join('.', '*.csv'))
list_of_csv_files.sort()
csvfile = list_of_csv_files[-1]  # last file in sorted list should be most recent

wb = Workbook()
ws = wb.active

# read in csvfile with pandas, then sort by first four columns
raw_data = pd.read_csv(csvfile)
sorted_data = raw_data.sort_values(by=['Supergroup','Group','Object name','Object ID number'])

# export data from pandas to openpyxl
for r in dataframe_to_rows(sorted_data, index=False, header=True):
    ws.append(r)

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDDDDD")

# freeze panes: header row, 3 leftmost columns (supergroup, group, name)
# ref: https://stackoverflow.com/questions/25588918/how-to-freeze-entire-header-row-in-openpyxl
pivot = ws['D2']
ws.freeze_panes = pivot

# set the column dimensions
ws.column_dimensions["A"].width = 25.0                  # Supergroup
ws.column_dimensions["B"].width = 32.0                  # Group
ws.column_dimensions["C"].width = 50.0                  # Item name
ws.column_dimensions["D"].width = 16.0                  # Object ID number
ws.column_dimensions["E"].width =  8.2                  # Quantity
ws.column_dimensions["F"].width = 21.14                 # Part number
ws.column_dimensions["G"].width = 23.57                 # Serial number
ws.column_dimensions["H"].width = 34.0                  # Donor...party
ws.column_dimensions["I"].width = 15.63                 # Acquisition method
ws.column_dimensions["J"].width = 27.0                  # Object lot ID
ws.column_dimensions["K"].width =  9.86                 # DGS paid
ws.column_dimensions["L"].width = 20.0                  # Transfer date
ws.column_dimensions["M"].width = 20.0                  # Delivery date
ws.column_dimensions["N"].width = 70.0                  # Current location
ws.column_dimensions["O"].width = 12.43                 # Accessioned?
ws.column_dimensions["P"].width = 10.0                  # ITAR status
ws.column_dimensions["Q"].width = 14.86                 # ITAR code
ws.column_dimensions["R"].width = 19.14                 # Owned by/loaned to
ws.column_dimensions["S"].width = 15.75                 # Item value
ws.column_dimensions["T"].width = 14.29                 # Insurance value
ws.column_dimensions["U"].width = 20.0                  # Insurance method
ws.column_dimensions["V"].width = 99.0                  # Notes

# trim off the .csv extension and save as .xlsx file
xlsxfile = csvfile[:-4] + '.xlsx'
wb.save(xlsxfile)

# copy this latest xlsx file to Dropbox
copy(xlsxfile, '/home/perry/Dropbox/CSC In-house/Artifact Documents/!Donations and Acquisitions')
